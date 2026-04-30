"""File-type-agnostic plain-text extraction for uploaded documents.

Used by the SoW field-extraction pipeline to produce a single string the
ML service can analyze, regardless of the source file's format. Also
called by the existing ``POST /api/sow/{id}/parse`` endpoint so PDF and
DOCX text extraction lives in one place.

Supported extensions:

* ``.pdf``  — via ``pypdf`` (the maintained successor to PyPDF2)
* ``.docx`` — stdlib zipfile + ElementTree (no external library)
* ``.csv``  — stdlib ``csv``, formatted as a markdown-style table
* ``.xlsx`` — via ``openpyxl``, one labeled markdown-style table per sheet

Anything else raises :class:`UnsupportedFileTypeError` so callers can
return a 415 instead of 500. CSV and XLSX iteration is capped at 1000
rows to keep prompts inside the LLM context window.
"""

from __future__ import annotations

import csv
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

EXTRACTABLE_EXTENSIONS = frozenset({".pdf", ".docx", ".csv", ".xlsx"})

# Cap rows we read from CSV/XLSX so a 100k-row staffing plan can't blow
# up the LLM prompt or backend memory.
_MAX_ROWS = 1000


class UnsupportedFileTypeError(ValueError):
    """Raised when a file extension has no registered extractor."""


def extract_text_pdf(file_path: str) -> str:
    """Extract text from a PDF using ``pypdf``."""
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def extract_text_docx(file_path: str) -> str:
    """Extract text from a DOCX without external dependencies.

    Reads ``word/document.xml`` from the .docx ZIP and pulls every ``w:t``
    text node, preserving paragraph boundaries via newlines.
    """
    try:
        with zipfile.ZipFile(file_path) as z:
            xml = z.read("word/document.xml")
    except (zipfile.BadZipFile, KeyError):
        return ""

    try:
        tree = ET.fromstring(xml)
    except ET.ParseError:
        return ""

    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs: list[str] = []
    for p in tree.findall(".//w:p", ns):
        texts = [t.text for t in p.findall(".//w:t", ns) if t.text]
        if texts:
            paragraphs.append("".join(texts))
    return "\n".join(paragraphs)


def extract_text_csv(file_path: str) -> str:
    """Read a CSV and format it as a markdown-style table for LLM context.

    Empty files return ``""``. The first row is treated as the header. We
    cap reading at ``_MAX_ROWS`` rows to keep the prompt bounded.
    """
    rows: list[list[str]] = []
    with open(file_path, encoding="utf-8-sig", newline="") as f:
        for row in csv.reader(f):
            rows.append([(cell or "").strip() for cell in row])
            if len(rows) >= _MAX_ROWS:
                break

    if not rows:
        return ""

    width = max(len(r) for r in rows)
    # Pad ragged rows so the markdown table renders consistently.
    padded = [r + [""] * (width - len(r)) for r in rows]
    header = " | ".join(padded[0])
    sep = " | ".join(["---"] * width)
    body_lines = [" | ".join(r) for r in padded[1:]]
    return "\n".join([header, sep, *body_lines])


def extract_text_xlsx(file_path: str) -> str:
    """Read an XLSX and format each sheet as a labeled markdown table.

    Uses ``openpyxl`` in read-only / data-only mode so formulas resolve to
    their last-cached values and we don't load the whole workbook into
    memory. Empty sheets are skipped silently.
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        sheet_rows: list[list[str]] = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= _MAX_ROWS:
                break
            sheet_rows.append(["" if c is None else str(c).strip() for c in row])

        # Drop trailing all-empty rows (very common in XLSX exports).
        while sheet_rows and not any(sheet_rows[-1]):
            sheet_rows.pop()
        if not sheet_rows:
            continue

        # Trim trailing all-empty columns so the table doesn't sprout
        # phantom columns from a worksheet that had wider formatting.
        width = max(len(r) for r in sheet_rows)
        for col in range(width - 1, -1, -1):
            if any(col < len(r) and r[col] for r in sheet_rows):
                width = col + 1
                break
        else:
            width = 0
        if width == 0:
            continue

        padded = [r[:width] + [""] * (width - len(r[:width])) for r in sheet_rows]
        header = " | ".join(padded[0])
        sep = " | ".join(["---"] * width)
        body_lines = [" | ".join(r) for r in padded[1:]]
        parts.append(f"## Sheet: {sheet.title}")
        parts.append(header)
        parts.append(sep)
        parts.extend(body_lines)
    return "\n".join(parts)


def extract_text_from_file(file_path: str) -> str:
    """Dispatch to the right extractor by extension.

    Raises :class:`UnsupportedFileTypeError` for any extension not in
    :data:`EXTRACTABLE_EXTENSIONS`.
    """
    ext = Path(file_path).suffix.lower()
    if ext == ".pdf":
        return extract_text_pdf(file_path)
    if ext == ".docx":
        return extract_text_docx(file_path)
    if ext == ".csv":
        return extract_text_csv(file_path)
    if ext == ".xlsx":
        return extract_text_xlsx(file_path)
    raise UnsupportedFileTypeError(f"Unsupported file type for extraction: {ext}")
